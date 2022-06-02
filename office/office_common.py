# 处理 office 相关文件
from ctypes import wstring_at
import string

from openpyxl import Workbook, load_workbook


class FixExcel:

  base_url = 'python/office/testResult/'  ##### --测试 office 文件存储位置--

  def __init__(self):
    pass

  def open_excel(self, url):
    load_workbook(url)
  
  def createExcel(self, name: string = 'sample'):
    wb = Workbook()
    ws = wb.active

    wb.save(f"{self.base_url}{name}.xlsx")

  def openExcel():
    wb = load_workbook(f'python/office/testResult/sample.xlsx')
    ws = wb.active

    ws1 = wb.create_sheet('sheet')
    ws1.title = "piplist"


# openExcel()
createExcel()